from AWSIoTPythonSDK.MQTTLib import AWSIoTMQTTClient
import cv2
import numpy as np
import face_recognition
import json
import requests
import datetime
import timeit
import time
import openpyxl
from openpyxl.styles import PatternFill

"""
1- Active
2- False positive double frame
3- Frame 
4- Duplicate
"""
writeSheet = 'attendance.xlsx'
wb = openpyxl.Workbook()
wb.save(writeSheet)
excel_workbook_write = openpyxl.load_workbook(writeSheet)
sheetN_write = excel_workbook_write.sheetnames
excel_worksheet_write = excel_workbook_write[sheetN_write[0]]
excel_worksheet_write.cell(row=1, column=1).value = "ID"
excel_worksheet_write.cell(row=1, column=2).value = "Name"
excel_worksheet_write.cell(row=1, column=3).value = "Time"
for z in range(1, 4):
    excel_worksheet_write.cell(row=1, column=z).fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00',fill_type="lightGrid")
excel_worksheet_write.auto_filter.ref = excel_worksheet_write.dimensions
excel_workbook_write.save(writeSheet)
ENDPOINT_URL = "a18vcw5xdqad81-ats.iot.me-south-1.amazonaws.com"
CA_PATH = "AmazonRootCA1.pem"
PRIVATE_KEY_PATH = "private.pem.key"
CERTIFICATE_PATH = "certificate.pem.crt"
cam_id = 7

start = ""
duplicate_array = []
first_frame = False
duplicateExcel=0

def receive(self, params, packet):
	global first_frame
	global duplicate_array
	global duplicateExcel



	first_frame = not first_frame
	response_str = (packet.payload).decode("utf-8")
	reasponse_json = json.loads(response_str)
	user_information = reasponse_json["payload"]
	if first_frame:
		duplicate_array.append(user_information)
	else:
		stop = timeit.default_timer()
		if user_information in duplicate_array:
			#for information in user_information:
				#print(information)
				

			for y in range(0,len(user_information)):
				try:
					for x in range(1,excel_worksheet_write.max_row+1):
						if user_information[y]["User_id"] == excel_worksheet_write.cell(row=x, column=1).value:
							duplicateExcel = 1
							break

					if (duplicateExcel != 1 and user_information[y]["User_id"] != "Unknown"):
						excel_worksheet_write.cell(row=excel_worksheet_write.max_row+1, column=1).value = user_information[y]["User_id"]
						excel_worksheet_write.cell(row=excel_worksheet_write.max_row, column=2).value = user_information[y]["Name"]
						excel_worksheet_write.cell(row=excel_worksheet_write.max_row, column=3).value = user_information[y]["Time"]
						print(user_information[y]["Name"] + " added to the excel sheet ")
					duplicateExcel=0
					excel_workbook_write.save(writeSheet)
	
				except PermissionError:
					print("Exception")


		duplicate_array = []
		#print("Req Time ", stop - start)


def send_to_cloud(encodings_str, cam_id):
    
	time = datetime.datetime.now()

	days = {"Saturday": "S", "Sunday": "U", "Monday": "M", "Tuesday": "T", "Wednesday": "W", "Thursday": "R",
	        "Friday": "F"}

	json_request = {
        "multiValueQueryStringParameters": {"face_encodings": encodings_str},
        "queryStringParameters":
        {"type":"camera","cam_id": cam_id,"time": time.strftime("%H:%M"),"day": days[time.strftime("%A")]}
        }

	test_request = {"multiValueQueryStringParameters": {"face_encodings": encodings_str},
	                "queryStringParameters": {"type":"camera","cam_id": cam_id, "time": "09:30", "day": "m"}}
	try:

		myMQTTClient.publish("kfupm/request", json.dumps(json_request), 0)
	except Exception:
		print(" ")


def encodings2str(encodings):
	encodings_str = []

	for encoding in encodings:
		# Convert numpy array to python array
		encoding_array = encoding.tolist()

		# Convert python array to string and append it to the face_encodings_str
		encodings_str.append(" ".join(str(e) for e in encoding_array))

	return encodings_str


myMQTTClient = AWSIoTMQTTClient(
	"SeniorClientID")  # random key, if another connection using the same key is opened the previous one is auto closed by AWS IOT
myMQTTClient.configureEndpoint(ENDPOINT_URL, 8883)
myMQTTClient.configureCredentials(CA_PATH, PRIVATE_KEY_PATH, CERTIFICATE_PATH)
myMQTTClient.configureOfflinePublishQueueing(-1)  # Infinite offline Publish queueing
myMQTTClient.configureDrainingFrequency(2)  # Draining: 2 Hz
myMQTTClient.configureConnectDisconnectTimeout(10)  # 10 sec
myMQTTClient.configureMQTTOperationTimeout(5)  # 5 sec
myMQTTClient.connect()
myMQTTClient.subscribe("{}/response".format(cam_id), 1, receive)

# open the camera
video_capture = cv2.VideoCapture(0)
video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
# Initialize some variables

process_this_frame = True

while True:
	# Grab a single frame of video

	ret, frame = video_capture.read()

	# Resize frame of video to 1/4 size for faster face recognition processing
	small_frame = cv2.resize(frame, (0, 0), None, 0.25, 0.25)

	# Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
	rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
	gray_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2GRAY)
	laplacian_var = cv2.Laplacian(gray_small_frame, cv2.CV_64F).var()

	if laplacian_var < 15:
		# print("bad")
		continue
	# Only process every other frame of video to save time
	if process_this_frame:
		face_locations = []
		face_encodings = []
		face_names = []
		# Find all the faces and face encodings in the current frame of video
		startBig = timeit.default_timer()
		face_locations = face_recognition.face_locations(rgb_small_frame)
		face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
		stopBig = timeit.default_timer()

		# Convert face encodings to str to ease sending it to the cloud
		if len(face_encodings) != 0:
			face_encodings_str = encodings2str(face_encodings)

			send_to_cloud(face_encodings_str, cam_id)
			#print("Process time: ", stopBig - startBig)
			if first_frame == False:
				start = timeit.default_timer()

	# Send the frame encodings do the cloud

	process_this_frame = not process_this_frame

	# Display the results

	# Display the resulting image
	cv2.imshow('Video', frame)

	# Hit 'q' on the keyboard to quit!
	if cv2.waitKey(1) & 0xFF == ord('q'):
		break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()

